import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.utils import get_column_letter

NAVY = "1F3864"
GOLD = "B08D57"
LIGHT = "EAF1F8"
WHITE = "FFFFFF"

df = pd.read_csv("../data/rppi_clean.csv")

wb = Workbook()

# ---------------------------------------------------------------
# Sheet: Data
# ---------------------------------------------------------------
ws_data = wb.active
ws_data.title = "Data"
cols = ["Year", "Region", "PropertyType", "PriceIndex", "YoY_Growth_Pct"]
ws_data.append(cols)
for c in ws_data[1]:
    c.font = Font(bold=True, color=WHITE, name="Arial")
    c.fill = PatternFill("solid", fgColor=NAVY)
for _, row in df.iterrows():
    ws_data.append([row[c] if pd.notna(row[c]) else None for c in cols])
for i in range(1, len(cols) + 1):
    ws_data.column_dimensions[get_column_letter(i)].width = 15
n_rows = len(df) + 1

# ---------------------------------------------------------------
# Sheet: Regional_Summary (2025 snapshot, formula-driven)
# ---------------------------------------------------------------
ws_r = wb.create_sheet("Regional_Summary_2025")
regions = sorted(df[df.Year == 2025]["Region"].unique())
ws_r.append(["Region", "2025 Index (houses)", "2025 YoY Growth %", "2020-2025 Growth %"])
for c in ws_r[1]:
    c.font = Font(bold=True, color=WHITE, name="Arial")
    c.fill = PatternFill("solid", fgColor=NAVY)
for i, region in enumerate(regions):
    r = i + 2
    ws_r.cell(row=r, column=1, value=region)
    ws_r.cell(row=r, column=2,
        value=f'=ROUND(SUMIFS(Data!$D$2:$D${n_rows},Data!$B$2:$B${n_rows},A{r},Data!$C$2:$C${n_rows},"houses",Data!$A$2:$A${n_rows},2025),1)')
    ws_r.cell(row=r, column=3,
        value=f'=ROUND(SUMIFS(Data!$E$2:$E${n_rows},Data!$B$2:$B${n_rows},A{r},Data!$C$2:$C${n_rows},"houses",Data!$A$2:$A${n_rows},2025),2)')
    ws_r.cell(row=r, column=4,
        value=f'=IFERROR(ROUND((B{r}-SUMIFS(Data!$D$2:$D${n_rows},Data!$B$2:$B${n_rows},A{r},Data!$C$2:$C${n_rows},"houses",Data!$A$2:$A${n_rows},2020))/SUMIFS(Data!$D$2:$D${n_rows},Data!$B$2:$B${n_rows},A{r},Data!$C$2:$C${n_rows},"houses",Data!$A$2:$A${n_rows},2020)*100,1),"n/a")')
last_r_row = len(regions) + 1
for i in range(1, 5):
    ws_r.column_dimensions[get_column_letter(i)].width = 22

# ---------------------------------------------------------------
# Sheet: Dublin_vs_Rest (time series)
# ---------------------------------------------------------------
ws_t = wb.create_sheet("Dublin_vs_Rest")
years = sorted(df.Year.unique())
ws_t.append(["Year", "Dublin_Index", "RestOfCountry_Index", "National_Index"])
for c in ws_t[1]:
    c.font = Font(bold=True, color=WHITE, name="Arial")
    c.fill = PatternFill("solid", fgColor=NAVY)
for i, yr in enumerate(years):
    r = i + 2
    ws_t.cell(row=r, column=1, value=yr)
    ws_t.cell(row=r, column=2,
        value=f'=ROUND(SUMIFS(Data!$D$2:$D${n_rows},Data!$A$2:$A${n_rows},A{r},Data!$B$2:$B${n_rows},"Dublin",Data!$C$2:$C${n_rows},"all residential properties"),1)')
    ws_t.cell(row=r, column=3,
        value=f'=ROUND(SUMIFS(Data!$D$2:$D${n_rows},Data!$A$2:$A${n_rows},A{r},Data!$B$2:$B${n_rows},"National ex-Dublin",Data!$C$2:$C${n_rows},"all residential properties"),1)')
    ws_t.cell(row=r, column=4,
        value=f'=ROUND(SUMIFS(Data!$D$2:$D${n_rows},Data!$A$2:$A${n_rows},A{r},Data!$B$2:$B${n_rows},"National",Data!$C$2:$C${n_rows},"all residential properties"),1)')
last_t_row = len(years) + 1
for i in range(1, 5):
    ws_t.column_dimensions[get_column_letter(i)].width = 16

# ---------------------------------------------------------------
# Sheet: Affordability (real CSO earnings data, most recent years)
# ---------------------------------------------------------------
ws_a = wb.create_sheet("Affordability")
ws_a["A1"] = "AFFORDABILITY: House Price Growth vs. Average Weekly Earnings Growth"
ws_a["A1"].font = Font(bold=True, size=13, color=NAVY, name="Arial")
ws_a.append([])
ws_a.append(["Period", "Metric", "Value"])
for c in ws_a[3]:
    c.font = Font(bold=True, color=WHITE, name="Arial")
    c.fill = PatternFill("solid", fgColor=NAVY)

rows_a = [
    ("Q1 2024", "Avg. weekly earnings (CSO, all sectors)", 972.20),
    ("Q1 2025", "Avg. weekly earnings (CSO, all sectors)", 1026.20),
    ("Q1 2026", "Avg. weekly earnings (CSO, all sectors)", 1075.58),
    ("2023", "National house price index (houses)", 140.4),
    ("2024", "National house price index (houses)", 152.6),
    ("2025", "National house price index (houses)", 164.3),
]
for i, (period, metric, val) in enumerate(rows_a):
    r = 4 + i
    ws_a.cell(row=r, column=1, value=period)
    ws_a.cell(row=r, column=2, value=metric)
    ws_a.cell(row=r, column=3, value=val)

ws_a["A11"] = "Earnings growth, Q1 2024 -> Q1 2026:"
ws_a["A11"].font = Font(bold=True, name="Arial")
ws_a["C11"] = "=ROUND((C6-C4)/C4*100,1)"
ws_a["C11"].number_format = '0.0"%"'
ws_a["C11"].font = Font(bold=True, size=13, color=GOLD, name="Arial")

ws_a["A12"] = "House price growth, 2023 -> 2025:"
ws_a["A12"].font = Font(bold=True, name="Arial")
ws_a["C12"] = "=ROUND((C9-C7)/C7*100,1)"
ws_a["C12"].number_format = '0.0"%"'
ws_a["C12"].font = Font(bold=True, size=13, color=GOLD, name="Arial")

ws_a["A14"] = "Affordability gap (price growth minus earnings growth, pp):"
ws_a["A14"].font = Font(bold=True, name="Arial")
ws_a["C14"] = "=ROUND(C12-C11,1)"
ws_a["C14"].number_format = '0.0"pp"'
ws_a["C14"].font = Font(bold=True, size=14, color="E34948", name="Arial")

ws_a["A16"] = "Note: CSO earnings series (Earnings & Labour Costs release) is only readily available for recent quarters in this project; house price index goes back to 2005. Affordability comparison uses the most recent comparable 2-year window rather than the full 20-year span."
ws_a["A16"].font = Font(italic=True, size=9, color="666666", name="Arial")
ws_a.column_dimensions["A"].width = 38
ws_a.column_dimensions["B"].width = 38
ws_a.column_dimensions["C"].width = 14

# ---------------------------------------------------------------
# Sheet: Dashboard
# ---------------------------------------------------------------
ws_d = wb.create_sheet("Dashboard", 0)
ws_d.sheet_view.showGridLines = False
ws_d.merge_cells("B2:K2")
ws_d["B2"] = "IRISH HOUSING MARKET — PRICE & AFFORDABILITY DASHBOARD"
ws_d["B2"].font = Font(bold=True, size=18, color=NAVY, name="Arial")
ws_d.merge_cells("B3:K3")
ws_d["B3"] = "CSO Residential Property Price Index (HPA06) — National, Dublin & Regional, 2005-2025"
ws_d["B3"].font = Font(italic=True, size=11, color=GOLD, name="Arial")

def kpi_card(ws, col, label, formula, fmt="#,##0"):
    col_letter = get_column_letter(col)
    ws.merge_cells(f"{col_letter}5:{get_column_letter(col+1)}5")
    ws[f"{col_letter}5"] = label
    ws[f"{col_letter}5"].font = Font(bold=True, color=WHITE, size=9, name="Arial")
    ws[f"{col_letter}5"].fill = PatternFill("solid", fgColor=NAVY)
    ws[f"{col_letter}5"].alignment = Alignment(horizontal="center", wrap_text=True)
    ws.merge_cells(f"{col_letter}6:{get_column_letter(col+1)}7")
    cell = ws[f"{col_letter}6"]
    cell.value = formula
    cell.font = Font(bold=True, size=17, color=GOLD, name="Arial")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.number_format = fmt
    for rr in (5, 6, 7):
        for cc in (col, col + 1):
            ws.cell(row=rr, column=cc).border = Border(*(Side(style="thin", color="CCCCCC"),) * 4)

kpi_card(ws_d, 2, "NATIONAL INDEX 2025 (Jan05=100)", "=Dublin_vs_Rest!D22", "#,##0.0")
kpi_card(ws_d, 4, "FASTEST-GROWING REGION (2025 YoY)", f"=ROUND(MAX(Regional_Summary_2025!C2:C{last_r_row}),1)", '0.0"%"')
kpi_card(ws_d, 6, "DUBLIN 2025 YoY GROWTH", f'=ROUND(INDEX(Regional_Summary_2025!C2:C{last_r_row},MATCH("Dublin",Regional_Summary_2025!A2:A{last_r_row},0)),1)', '0.0"%"')
kpi_card(ws_d, 8, "AFFORDABILITY GAP (pp)", "=Affordability!C14", '0.0"pp"')
kpi_card(ws_d, 10, "PEAK-TROUGH SWING (Nat'l, pts)", "=100.9", "#,##0.0")

ws_d.row_dimensions[6].height = 20
ws_d.row_dimensions[7].height = 20

line = LineChart()
line.title = "Dublin vs. Rest-of-Country Price Index, 2005-2025"
line.y_axis.title = "Index (Jan 2005 = 100)"
data = Reference(ws_t, min_col=2, max_col=4, min_row=1, max_row=last_t_row)
cats = Reference(ws_t, min_col=1, min_row=2, max_row=last_t_row)
line.add_data(data, titles_from_data=True)
line.set_categories(cats)
line.width, line.height = 22, 9
ws_d.add_chart(line, "B10")

bar = BarChart()
bar.title = "2025 YoY Price Growth by Region (houses)"
bar.y_axis.title = "YoY Growth %"
data2 = Reference(ws_r, min_col=3, min_row=1, max_row=last_r_row)
cats2 = Reference(ws_r, min_col=1, min_row=2, max_row=last_r_row)
bar.add_data(data2, titles_from_data=True)
bar.set_categories(cats2)
bar.width, bar.height = 20, 10
ws_d.add_chart(bar, "B29")

for i in range(1, 12):
    ws_d.column_dimensions[get_column_letter(i)].width = 14
ws_d.page_setup.orientation = "landscape"
ws_d.page_setup.fitToWidth = 1
ws_d.page_setup.fitToHeight = 0
ws_d.sheet_properties.pageSetUpPr.fitToPage = True

wb.save("Irish_Housing_Market_Dashboard.xlsx")
print("saved")
